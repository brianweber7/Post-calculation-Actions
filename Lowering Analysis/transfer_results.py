'''
Created by: Brian Weber
Project: 1601 DOF Dynamic Lowering Analysis
Date Created: 1/8/2015

Description: This script is to be used to transfer results from
text files to an excel spreadsheet.
'''

import glob, os
import re
import openpyxl


def tryint(s):
    try:
        return int(s)
    except:
        return s

def alphanum_key(s):
    """ Turn a string into a list of string and number chunks.
        "z23a" -> ["z", 23, "a"]
    """
    return [ tryint(c) for c in re.split('([0-9]+)', s) ]

def sort_nicely(l):
    """ Sort the given list in the way that humans expect.
    """
    l.sort(key=alphanum_key)

files = []
for file in glob.glob("*.txt"):
	if 'DOF' in file:
		pass
	else:
		files.append(file)
files.sort(key=alphanum_key)

# create excel file
wb = openpyxl.Workbook()
wb.create_sheet(index=0, title='Tensions')
wb.create_sheet(index=1, title='Plate Slam Force')
wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
tension_sheet = wb.get_sheet_by_name('Tensions')
Plate_SF_sheet = wb.get_sheet_by_name('Plate Slam Force')

# Write data to an Excel file
for file in files:
	file_split = file.split('.')
	file_name = file_split[0]
	with open(file, 'r') as f:
		lines = f.readlines()
	f.close()

	tension_line = [lines[1]]
	for tension in tension_line:
		tensions = tension.split(', ')
	tensions.insert(0, file_name)

	slamming_force_line = [lines[3]]
	for slamming_force in slamming_force_line:
		slamming_forces = slamming_force.split(', ')
	slamming_forces.insert(0, file_name)
	
	tension_sheet.append(tensions)
	Plate_SF_sheet.append(slamming_forces)

# Save file
wb.save('transfer_results.xlsx')

